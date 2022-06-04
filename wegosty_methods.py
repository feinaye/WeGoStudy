from selenium import webdriver
import wegosty_locators as locators
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from time import sleep
import datetime
from selenium.webdriver.common.by import By
import sys
import random
from selenium.webdriver.common.keys import Keys
from  openpyxl import Workbook
from openpyxl.styles import fonts

s = Service(executable_path='chromedriver.exe')
driver = webdriver.Chrome(service=s)


def setUp():
    print(f'launch {locators.app}')
    print('___________________________________________')
    driver.maximize_window()
    driver.implicitly_wait(30)
    driver.get(locators.wegosty_url)
    if driver.current_url == locators.wegosty_homepage_url:
        print(f' WeGoStudy URL: {driver.current_url}')
        print(f'__________________Test started successfully at {datetime.datetime.now()}__________________')

def login():
    # driver.find_element(By.CSS_SELECTOR, '.toast-message').click()
    # sleep(0.3)
    driver.find_element(By.LINK_TEXT, 'LOGIN').click()
    sleep(0.3)
    driver.find_element(By.ID, 'user_email').send_keys('sufenyu2012@hotmail.com')
    sleep(0.3)
    driver.find_element(By.ID, 'user_password').send_keys(f'Password2')
    sleep(0.3)
    # driver.find_element(By.ID, 'user_password').send_keys(f'123cctb{Keys.ENTER}')
    driver.find_element(By.CSS_SELECTOR, 'input[value="SIGN IN"]').click()
    # driver.find_element(By.CSS_SELECTOR, 'input[type="submit"]').click()
    # driver.find_element(By.XPATH, "//input[@value='SIGN IN']").click()
    assert driver.find_element(By.CSS_SELECTOR, 'img[alt="Partner"]').is_displayed()
    print(f'________Signed in successfully! at {datetime.datetime.now()}_________')


def create_new_student():
    driver.find_element(By.LINK_TEXT, 'My WeGoStudy').click()
    sleep(1)
    driver.find_element(By.LINK_TEXT, 'Students').click()
    sleep(1)
    driver.find_element(By.LINK_TEXT, 'Create New Student').click()
    sleep(1)
    # __________________Personal_Information______________________________
    driver.find_element(By.ID, 'user_student_detail_attributes_first_name').send_keys(locators.first_name)
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_last_name').send_keys(locators.last_name)
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_preferred_name').send_keys(locators.first_name)
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_birth_date').send_keys(Keys.ENTER)
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_birth_date').clear()
    sleep(0.3)
    driver.find_element(By.CSS_SELECTOR, 'input[placeholder="Date of Birth on passport"]').send_keys('20000202')
    # driver.find_element(By.ID, 'user_student_detail_attributes_birth_date').send_keys(locators.date_of_birth)
    # driver.find_element(By.ID, 'user_student_detail_attributes_birth_date').send_keys(f'{locators.date_of_birth}{Keys.ENTER}')
    sleep(0.3)
    driver.find_element(By.ID, 'phone_number').send_keys(locators.phone_number)
    sleep(0.3)
    driver.find_element(By.CSS_SELECTOR, 'span[title="Country of citizenship"]').click()
    sleep(0.3)
    driver.find_element(By.CSS_SELECTOR, 'input[role="searchbox"]').send_keys(f'Cameroon{Keys.ENTER}')
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_passport_number').send_keys(locators.passport_number)
    sleep(0.3)
    # __________________Contact Information______________________________
    driver.find_element(By.ID, 'user_student_detail_attributes_address_attributes_mailing_address').send_keys(
        locators.mailing_address)
    sleep(0.3)
    driver.find_element(By.LINK_TEXT, 'Country').click()
    sleep(0.3)
    driver.find_element(By.XPATH,
                        '//*[@id="user_student_detail_attributes_address_attributes_country_chosen"]/div/div/input').send_keys(
        f'Canada{Keys.ENTER}')
    # driver.find_element(By.XPATH, '//*[@id="user_student_detail_attributes_address_attributes_country_chosen"]/div/ul/li[40]').click()
    sleep(0.6)
    driver.find_element(By.LINK_TEXT, 'Province/State').click()
    sleep(0.6)
    driver.find_element(By.XPATH,
                        '//*[@id="user_student_detail_attributes_address_attributes_state_chosen"]/div/ul/li[3]').click()
    sleep(0.6)
    driver.find_element(By.LINK_TEXT, 'City').click()
    sleep(0.6)
    driver.find_element(By.XPATH,
                        '//*[@id="user_student_detail_attributes_address_attributes_city_chosen"]/div/ul/li[30]').click()
    # driver.find_element(By.XPATH, '//*[@id="user_student_detail_attributes_address_attributes_city_chosen"]/a/span').click()
    sleep(0.6)
    driver.find_element(By.ID, 'user_student_detail_attributes_address_attributes_zip_code').send_keys(
        locators.postal_code)
    sleep(0.6)
    driver.find_element(By.ID, 'user_email').send_keys(locators.user_email)
    sleep(0.6)
    # __________________Education Information______________________________
    driver.find_element(By.LINK_TEXT, 'Credentials').click()
    sleep(0.6)
    driver.find_element(By.XPATH,
                        '//*[@id="user_student_detail_attributes_user_educations_attributes_0_credentials_chosen"]/div/ul/li[3]').click()
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_user_educations_attributes_0_school_name').send_keys(
        'CCTB')
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_user_educations_attributes_0_program').send_keys('SQTA')
    sleep(0.3)
    driver.find_element(By.LINK_TEXT, 'GPA Scale').click()
    sleep(0.3)
    driver.find_element(By.XPATH,
                        '//div[@id="user_student_detail_attributes_user_educations_attributes_0_gpa_scale_chosen"]/div/ul/li[5]').click()
    sleep(0.3)
    # __________________________________________
    driver.find_element(By.ID, 'user_student_detail_attributes_user_educations_attributes_0_gpa').send_keys(f'4.0')
    sleep(0.3)
    driver.find_element(By.XPATH, '//input[@value="Save"]').submit()
    # __________________________________________
    print(f'The new student {locators.first_name} {locators.last_name} is created.')
    print(
        f'Personal Information: Date Of Birth - {locators.date_of_birth}, Passport Number - {locators.passport_number}, Citizenship - Canada, Phone Number - {locators.phone_number}')
    print(
        f'Contact Information: Mailing Address - {locators.mailing_address}, Country - Canada, State - BC, City -  Burnaby, Postal Code - {locators.postal_code}')
    # print(f'Credential -, School Name -, Program -, GPA Scale -, Scored GPA - ')
    print('___________________________________________')


def create_application():
    driver.find_element(By.LINK_TEXT, 'My WeGoStudy').click()
    sleep(0.6)
    driver.find_element(By.LINK_TEXT, 'Application').click()
    sleep(0.6)
    driver.find_element(By.LINK_TEXT, 'Create Application').click()
    sleep(0.6)
    driver.find_element(By.ID, 'select2-admission_institute_detail_id-container').click()
    sleep(0.6)
    # driver.find_element(By.XPATH, '//div[@id="admission_institute_program_id_chosen"]/a/span').click()
    # sleep(0.6)


def search_student():
    from openpyxl import Workbook
    import openpyxl as O
    Exel_file = "C:\\Users\\ks\\Desktop\\OXANA\\QA\\wgsty\\WEG_test_result.xlsx"
    wb = O.load_workbook(Exel_file)
    ws = wb.active

    driver.find_element(By.LINK_TEXT, 'My WeGoStudy').click()
    sleep(0.6)
    driver.find_element(By.LINK_TEXT, 'Students').click()
    sleep(0.6)
    driver.find_element(By.ID, 'search').send_keys(locators.first_name, locators.last_name)
    # driver.find_element(By.ID, 'search').send_keys(locators.first_name, locators.last_name)
    driver.find_element(By.ID, 'search').send_keys(ws['A5'].value, ' ', ws['B5'].value)
    sleep(0.6)
    driver.find_element(By.NAME, 'commit').click()


def tearDown():
    if driver is not None:
        print(f'_________________Test finished successfully at {datetime.datetime.now()}_________________')
        sleep(2)
        driver.close()
        driver.quit()
        logger('delleted')


def logger(action: object):
    old_instance = sys.stdout
    log_file = open('message.log', 'a')
    sys.stdout = log_file
    # print(f'{locators.user_email}\t'
    #       f'{locators.first_name}\t'
    #       f'{locators.last_name}\t'
    #       f'{locators.mailing_address}\t'
    #       f'{locators.postal_code}\t'
    #       f'{datetime.phone_number()}\t'
    #       f'{datetime.datetime.now()}\t'
    #       f'{action}')
    sys.stdout = old_instance
    log_file.close()


def xlsx_data():
    from openpyxl import Workbook
    import openpyxl as O
    Exel_file = "C:\\Users\\Lenovo\\Desktop\\practicum_wgy\\WEG_test_result.xlsx"
    wb = O.load_workbook(Exel_file)
    ws = wb.active
    ws.append([f'{locators.first_name}', f'{locators.last_name}', f'{locators.passport_number}',
               f'{locators.mailing_address}',
               f'{locators.postal_code}', f'{locators.phone_number}', f'{locators.user_email}',
             f'{datetime.datetime.now()}',
               "test result"])
    wb.save(Exel_file)


def view_details():
        driver.find_element(By.XPATH, '//*[@id="navbar-nav"]/ul[1]/li[1]/a/span[contains(., "My WeGoStudy")]').click()
        sleep(0.6)
        driver.find_element(By.LINK_TEXT, 'Students').click()
        sleep(0.6)
        if driver.find_element(By.XPATH, '//*[@id="student_list"]/div[1]/div[3]/a[1][contains(., "View Details")]').is_enabled():
            print('View Details is clickable.')
            driver.find_element(By.XPATH, '//*[@id="student_list"]/div[1]/div[3]/a[1][contains(., "View Details")]').click()
            driver.find_element(By.XPATH, '/html/body/form/section/div/div/div[2]/div[1]/div/div[1]').is_displayed()
            print('Personal Information has been completed.')
            driver.find_element(By.XPATH, '/html/body/form/section/div/div/div[2]/div[2]/div/div[1]/div').is_displayed()
            print('Contact Information has been completed.')
            driver.find_element(By.XPATH, '/html/body/form/section/div/div/div[2]/div[3]/div/div[1]/h5').is_displayed()
            print('Education Information has been completed.')
            driver.find_element(By.XPATH, '/html/body/form/section/div/div/div[2]/div[4]/div/div[1]/div/div[2]/div/div/div/div/img').is_displayed()
            print('Documents have been uploaded.')

def view_applications():
    # View student application page
    driver.find_element(By.XPATH, '//*[@id="navbar-nav"]/ul[1]/li[1]/a/span[contains(., "My WeGoStudy")]').click()
    sleep(0.6)
    driver.find_element(By.LINK_TEXT, 'Students').click()
    sleep(0.6)
    if driver.find_element(By.XPATH, '//*[@id="student_list"]/div[1]/div[3]/a[3][contains(., "View Applications")]').is_enabled():
        sleep(0.6)
        driver.find_element(By.XPATH, '//*[@id="student_list"]/div[1]/div[3]/a[3][contains(., "View Applications")]').click()
        sleep(0.6)
        driver.find_element(By.ID, 'admission_list').is_displayed()
        sleep(0.6)
        print('Admission List has been displayed.')
        sleep(0.6)
        driver.find_element(By.ID, 'admission_list_first').is_enabled()
        sleep(0.6)
        driver.find_element(By.ID, 'admission_list_previous').is_enabled()
        sleep(0.6)
        driver.find_element(By.ID, 'admission_list_next').is_enabled()
        sleep(0.6)
        driver.find_element(By.ID, 'admission_list_last').is_enabled()
        sleep(0.6)
        print('Admission List Paginate is clickable.')
        sleep(0.6)
        driver.find_element(By.XPATH, '//*[@id="pay_for_application"][contains(., "Pay For Application")]').is_enabled()
        sleep(0.6)
        print('Pay for Application Button is clickable.')
        sleep(0.6)
        driver.find_element(By.XPATH, '//*[@id="send_message"][contains(., "Send Message")]').is_enabled()
        sleep(0.6)
        print('Send Message Button is clickable.')
        sleep(0.6)
        driver.find_element(By.XPATH, '//*[@id="delete_applications"][contains(., "Delete Application(s)")]').is_enabled()
        sleep(0.6)
        print('Delete Application Button is clickable.')
        sleep(0.6)
        driver.find_element(By.XPATH, '//*[@id="admission_list"]/thead/tr/th[10][contains(., "Chat")]').is_enabled()
        sleep(0.6)
        print('Chat button with site administrator is clickable.')
        sleep(0.6)
        driver.find_element(By.XPATH, '//*[@id="large_modal"]/div/div/div[2]/div[2]/button').click()
        sleep(0.6)


def send_message_student():
    driver.find_element(By.XPATH, '//*[@id="navbar-nav"]/ul[1]/li[1]/a/span[contains(., "My WeGoStudy")]').click()
    sleep(0.6)
    driver.find_element(By.LINK_TEXT, 'Students').click()
    sleep(0.6)
    driver.find_element(By.XPATH, '//*[@id="student_list"]/div[1]/div[3]/a[4][contains(., "Send Message")]').is_enabled()
    sleep(0.6)
    driver.find_element(By.XPATH, '//*[@id="student_list"]/div[1]/div[3]/a[4][contains(., "Send Message")]').click()
    sleep(2)
    driver.find_element(By.XPATH, '//*[@id="modal"]/div/div/div[1]/h4[contains(., "Send Message")]').is_displayed()
    sleep(0.6)
    print('Send Message to Student Page has been displayed.')
    sleep(0.6)
    driver.find_element(By.ID, 'message').send_keys(locators.message_student)
    sleep(0.25)
    driver.find_element(By.XPATH, '//*[@id="modal"]/div/div/div[2]/form/div[2]/input').click()
    sleep(0.25)
    print('Message sent successfully!')


def log_out():
    sleep(5)
    # driver.find_element(By.XPATH, '//div[@id="navbar-nav"]/ul[2]/li[2]/a/span').click()
    driver.find_element(By.XPATH, '//*[@id="navbar-nav"]/ul[2]/li[2]/a/span').click()
    sleep(1)
    driver.find_element(By.LINK_TEXT, 'Log out').click()
    print(f'________Singed out successfully! at {datetime.datetime.now()}_________')
    # sleep(0.8)
    # driver.find_element(By.XPATH, '//*[@id="navbar-nav"]/ul[2]/li[2]/a/span[contains(., "Sufen Yu")]').click()
    # sleep(0.3)
    # driver.find_element(By.CSS_SELECTOR, 'span[class="my-auto mr-2 pf-name"]').click()
    # sleep(0.3)
    # driver.find_element(By.XPATH, '//*[@id="navbar-nav"]/ul[2]/li[2]/div/a[4]').click()
    # if driver.find_element(By.XPATH, '//div[@id="toast-container"]').is_displayed():
    #     print(f'________Singed out successfully! at {datetime.datetime.now()}_________')
    # driver.find_element(By.XPATH, '//div[@id="toast-container"]').click()


setUp()
login()
# create_new_student()
# create_application()
view_details()
view_applications()
send_message_student()
# # search_student()
# # xlsx_data()
log_out()
tearDown()
